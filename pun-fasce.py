import argparse
import requests
import zipfile
import io
from openpyxl import load_workbook
from datetime import date
import holidays
from statistics import mean
import csv

def get_fascia(data, festivo, ora):
	#F1 = lu-ve 8-19
	#F2 = lu-ve 7-8, lu-ve 19-23, sa 7-23
	#F3 = lu-sa 0-7, lu-sa 23-24, do, festivi
	if festivo or (data.weekday() == 6):
		# Festivi e domeniche
		return 3
	elif (data.weekday() == 5):
		# Sabato
		if (ora >= 7) and (ora < 23):
			return 2
		else:
			return 3
	else:
		# Altri giorni della settimana
		if (ora == 7) or ((ora >= 19) and (ora < 23)):
			return 2
		elif (ora == 23) or ((ora >= 0) and (ora < 7)):
			return 3
	return 1

# Formatta la media come numero decimale con 6 decimali (ma arrotondato al quinto)
def fmt_mean(values):
    return format(round(mean(values), 5), '.6f')

def calc_f23(f2, f3):
	# Calcola la fascia F23 sulla base delle fasce F2 e F3
 	# La motivazione del calcolo è oscura, ma sembra corretta, vedere:
	# https://github.com/virtualdj/pun_sensor/issues/24#issuecomment-1829846806
	return format(round(mean(f2), 5) * .46 + round(mean(f3), 5) * .54, '.6f')


def parse_date_range(start_date_str, end_date_str):
    start_year, start_month = map(int, start_date_str.split('-'))
    end_year, end_month = map(int, end_date_str.split('-'))
    
    start_date = date(start_year, start_month, 1)
    
    if end_month == 12:
        end_date = date(end_year, 12, 31)
    else:    
        end_date = date(end_year, end_month+1, 1)
    
    
    if start_date > end_date:
        raise ValueError('La data di inizio è successiva alla data di fine.')
    return start_date, end_date

def download_data(year):
    if year < 2016 or year > date.today().year:
        raise SystemExit('ERRORE: anno non valido! Inserire un anno tra 2016 e l\'anno corrente.')
    
    url = f'https://www.mercatoelettrico.org/it-it/Home/Esiti/Elettricita/MGP/Statistiche/DatiStorici/moduleId/10874/controller/GmeDatiStoriciItem/action/DownloadFile?fileName=Anno{year}.zip'
    try:
        response = requests.get(url)
        response.raise_for_status()
        return zipfile.ZipFile(io.BytesIO(response.content))
    except requests.RequestException as e:
        raise SystemExit(f'ERRORE: sito web MercatoElettrico.org non disponibile! ({e})')
    except zipfile.BadZipFile:
        raise SystemExit('ERRORE: il file scaricato dal sito non è valido.')

def extract_excel_file(archive):
    for file in archive.filelist:
        if file.filename.lower().endswith('xlsx'):
            return archive.open(file.filename)
    raise SystemExit('ERRORE: nessun file Excel è stato trovato nel file scaricato dal sito.')

def load_excel_sheet(xlfile):
    try:
        workbook = load_workbook(xlfile)
        return workbook['Prezzi-Prices']
    except KeyError:
        raise SystemExit('ERRORE: file Excel "Prezzi" non trovato nel file scaricato dal sito.')
    except Exception:
        raise SystemExit('ERRORE: file Excel non valido nel file scaricato dal sito.')

def process_data(sheet, start_date, end_date):
    it_holidays = holidays.IT()
    prev_dat = ''
    prev_month = 0
    f1, f2, f3, monoorario = [], [], [], []
    output_data = [['Mese', 'MO (€/kWh)', 'F1 (€/kWh)', 'F2 (€/kWh)', 'F3 (€/kWh)', 'F23 (€/kWh)']]

    for row in range(2, sheet.max_row):
        if sheet.cell(row, 1).value is None:
            break

        # Estrae i valori delle celle dell'Excel
        dat = str(sheet.cell(row, 1).value) #YYYYMMDD
        ora = int(sheet.cell(row, 2).value) - 1 # 1..24
        prezzo = float(sheet.cell(row, 3).value) / 1000

        # Verifica se il giorno è cambiato
        if dat != prev_dat:
            
            # Converte la stringa giorno in data
            dat2 = date(int(dat[0:4]), int(dat[4:6]), int(dat[6:8]))
            
            festivo = dat2 in it_holidays            
            prev_dat = dat

        # Verifica se il mese è cambiato
        if start_date <= dat2 < end_date:
            if dat2.month != prev_month:
                
                # Nuovo mese	
                if prev_month > 0 and monoorario:
                    month_row = [
                        f'{prev_month:02d}/{start_date.year}',
                        fmt_mean(monoorario),
                        fmt_mean(f1),
                        fmt_mean(f2),
                        fmt_mean(f3),
                        calc_f23(f2, f3)
                    ]
                    output_data.append(month_row)

                prev_month = dat2.month
                f1.clear()
                f2.clear()
                f3.clear()
                monoorario.clear()

            fascia = get_fascia(dat2, festivo, ora)
            if fascia == 3:
                f3.append(prezzo)
            elif fascia == 2:
                f2.append(prezzo)
            elif fascia == 1:
                f1.append(prezzo)
            monoorario.append(prezzo)

    if monoorario:
        month_row = [
            f'{prev_month:02d}/{start_date.year}',
            fmt_mean(monoorario),
            fmt_mean(f1),
            fmt_mean(f2),
            fmt_mean(f3),
            calc_f23(f2, f3)
        ]
        output_data.append(month_row)

    return output_data

def save_to_csv(output_data, start_date, end_date):
    with open(f'output_{start_date}_{end_date}.csv', mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(output_data)
        
def parse_arguments():
    parser = argparse.ArgumentParser(description='Mostra i costi del PUN di un anno, diviso per mesi e fasce orarie.')
    parser.add_argument('start_date', type=str, help='Data di inizio in formato "YYYY-MM"')
    parser.add_argument('end_date', type=str, help='Data di fine in formato "YYYY-MM"')
    parser.add_argument('-csv', action='store_true', help='Esporta l\'output in formato CSV')
    return parser.parse_args()

def main():
    args = parse_arguments()
    
    try:
        start_date, end_date = parse_date_range(args.start_date, args.end_date)
    except Exception as e:
        raise SystemExit(f'ERRORE: formato dell\'intervallo di date non valido! ({e})')

    output_data = [['Mese', 'MO (€/kWh)', 'F1 (€/kWh)', 'F2 (€/kWh)', 'F3 (€/kWh)', 'F23 (€/kWh)']]
    
    for year in range(start_date.year, end_date.year + 1):
        print(f'Download {year} data...')
        archive = download_data(year)
        print(f'Extracting...')
        xlfile = extract_excel_file(archive)
        sheet = load_excel_sheet(xlfile)
        print(f'Processing...')
        yearly_data = process_data(sheet, start_date if year == start_date.year else date(year, 1, 1), end_date if year == end_date.year else date(year, 12, 31))
        output_data.extend(yearly_data[1:])  # Skip header row for subsequent years
    
    print()
    
    if args.csv:
        save_to_csv(output_data, start_date, end_date)
    else:
        for row in output_data:
            print(*row, sep='\t')

if __name__ == '__main__':
    main()